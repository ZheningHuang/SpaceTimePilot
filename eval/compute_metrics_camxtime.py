#!/usr/bin/env python3
"""
eval/compute_metrics_camxtime.py

Evaluate SpaceTimePilot moved-cam→moved-cam results against preprocessed GT videos.

Metrics: PSNR, SSIM, LPIPS (AlexNet), computed per frame then averaged per video.

Path conventions
----------------
  Predictions : {pred_root}/{mode}/{scene}_cam_00.mp4
  GT          : {gt_root}/{scene}/moving_{gt_pattern}.mp4

Mode → GT pattern mapping
--------------------------
  fixed_10  →  moving_bullettime
  normal    →  moving_forward
  reverse   →  moving_backward
  slowmo    →  moving_slowmo
  zigzag    →  moving_zigzag

Outputs (written to --output_dir)
----------------------------------
  metrics.json          full per-video / per-frame data
  per_video.csv         one row per (mode, scene)
  summary.csv           one row per mode (mean ± std)
  results.xlsx          Summary sheet + one sheet per mode
  psnr_overview.png     bar chart: all scenes × all modes
  ssim_overview.png
  lpips_overview.png
  summary_bar.png       3-panel summary with error bars

Usage (run from repo root):
    python eval/compute_metrics_camxtime.py \\
        --pred_root  results/moved_cam2moved_cam_extended \\
        --gt_root    CamxTime_eval/eval_gt_wan2.1_format \\
        --metadata   metadata.csv \\
        --output_dir results/camxtime_metrics
"""

import argparse
import csv
import json
import time
from pathlib import Path

import imageio.v2 as imageio
import lpips
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from skimage.metrics import peak_signal_noise_ratio as psnr_fn
from skimage.metrics import structural_similarity as ssim_fn
from tqdm import tqdm

# ── Mode definitions ──────────────────────────────────────────────────────────
MODES = ["fixed_10", "normal", "reverse", "slowmo", "zigzag"]

MODE_META = {
    # mode_key: (display_label, gt_pattern_suffix)
    "fixed_10": ("Bullet Time", "bullettime"),
    "normal":   ("Forward",     "forward"),
    "reverse":  ("Backward",    "backward"),
    "slowmo":   ("Slow Motion", "slowmo"),
    "zigzag":   ("Zigzag",      "zigzag"),
}

VIDEO_H, VIDEO_W = 480, 832
LPIPS_BATCH = 16

MODE_COLORS = {
    "fixed_10": "#5C85D6",
    "normal":   "#4CAF50",
    "reverse":  "#FF9800",
    "slowmo":   "#9C27B0",
    "zigzag":   "#F44336",
}

# ── Terminal helpers ──────────────────────────────────────────────────────────
_W = 72   # print width

def _sep(char="─"):
    print(char * _W)

def _banner(text, char="═"):
    pad = max(0, _W - len(text) - 4)
    print(f"{char*2} {text} {char * pad}")

def _table_row(cols, widths, sep="│"):
    parts = [f" {str(v):<{w}} " for v, w in zip(cols, widths)]
    print(sep + sep.join(parts) + sep)

def _table_header(cols, widths):
    _table_row(cols, widths)
    _sep("─")

def _running_stats_line(rows, label="running"):
    """Print a compact running-mean line from accumulated rows."""
    if not rows:
        return
    p = np.mean([r["psnr"]  for r in rows])
    s = np.mean([r["ssim"]  for r in rows])
    l = np.mean([r["lpips"] for r in rows])
    print(f"  ↳ {label:>12s}  PSNR {p:6.3f}  SSIM {s:.4f}  LPIPS {l:.4f}"
          f"  [{len(rows)} videos]")


# ── Video I/O ─────────────────────────────────────────────────────────────────
def load_frames(path: Path) -> np.ndarray:
    reader = imageio.get_reader(str(path))
    frames = []
    try:
        for frame in reader:
            img = np.asarray(frame)
            if img.shape[0] != VIDEO_H or img.shape[1] != VIDEO_W:
                img = np.array(Image.fromarray(img).resize((VIDEO_W, VIDEO_H), Image.BILINEAR))
            frames.append(img[:, :, :3])
    finally:
        reader.close()
    return np.stack(frames, axis=0)


def to_lpips_tensor(frames: np.ndarray, device: str) -> torch.Tensor:
    t = torch.from_numpy(frames).float() / 127.5 - 1.0
    return t.permute(0, 3, 1, 2).to(device)


# ── Per-video metrics ─────────────────────────────────────────────────────────
def compute_video_metrics(gt, pred, lpips_net, device):
    T = min(len(gt), len(pred))
    gt, pred = gt[:T], pred[:T]
    psnr_vals = [float(psnr_fn(g, p, data_range=255)) for g, p in zip(gt, pred)]
    ssim_vals = [float(ssim_fn(g, p, data_range=255, channel_axis=-1)) for g, p in zip(gt, pred)]
    gt_t, pred_t = to_lpips_tensor(gt, device), to_lpips_tensor(pred, device)
    lpips_vals = []
    with torch.no_grad():
        for i in range(0, T, LPIPS_BATCH):
            d = lpips_net(gt_t[i:i+LPIPS_BATCH], pred_t[i:i+LPIPS_BATCH])
            v = d.squeeze().cpu()
            lpips_vals.extend(v.tolist() if v.ndim > 0 else [float(v)])
    return {
        "psnr": float(np.mean(psnr_vals)), "ssim": float(np.mean(ssim_vals)),
        "lpips": float(np.mean(lpips_vals)), "num_frames": T,
        "psnr_per_frame": psnr_vals, "ssim_per_frame": ssim_vals,
        "lpips_per_frame": lpips_vals,
    }


# ── Visualisation ─────────────────────────────────────────────────────────────
METRIC_INFO = {
    "psnr":  dict(label="PSNR (dB)"),
    "ssim":  dict(label="SSIM"),
    "lpips": dict(label="LPIPS"),
}


# ── Excel ─────────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)
AVG_FILL = PatternFill("solid", fgColor="D9E1F2")
AVG_FONT = Font(bold=True)


def _header(ws, cols):
    for c, lbl in enumerate(cols, 1):
        cell = ws.cell(1, c, lbl)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def add_mode_sheet(wb, mode, rows):
    ws = wb.create_sheet(title=MODE_META[mode][0])
    _header(ws, ["Scene", "Frames", "PSNR (dB)", "SSIM", "LPIPS"])
    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["scene"]); ws.cell(r, 2, row["num_frames"])
        ws.cell(r, 3, round(row["psnr"], 3)); ws.cell(r, 4, round(row["ssim"], 4))
        ws.cell(r, 5, round(row["lpips"], 4))
    if rows:
        ar = len(rows) + 2
        ws.cell(ar, 1, "MEAN").font = AVG_FONT
        for col, key in [(3, "psnr"), (4, "ssim"), (5, "lpips")]:
            cell = ws.cell(ar, col, round(float(np.mean([r[key] for r in rows])), 4))
            cell.fill = AVG_FILL; cell.font = AVG_FONT
    ws.column_dimensions["A"].width = 16


def add_summary_sheet(wb, summary, modes):
    ws = wb.create_sheet("Summary", 0)
    _header(ws, ["Mode", "Videos", "PSNR mean", "PSNR std",
                 "SSIM mean", "SSIM std", "LPIPS mean", "LPIPS std"])
    for r, mode in enumerate(modes, 2):
        if mode not in summary:
            continue
        s = summary[mode]
        ws.cell(r, 1, MODE_META[mode][0]); ws.cell(r, 2, s["num_videos"])
        ws.cell(r, 3, round(s["psnr_mean"], 3));  ws.cell(r, 4, round(s["psnr_std"], 3))
        ws.cell(r, 5, round(s["ssim_mean"], 4));  ws.cell(r, 6, round(s["ssim_std"], 4))
        ws.cell(r, 7, round(s["lpips_mean"], 4)); ws.cell(r, 8, round(s["lpips_std"], 4))
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["A"].width = 16


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Compute PSNR/SSIM/LPIPS for SpaceTimePilot moved-cam results."
    )
    parser.add_argument("--pred_root",  default="results/moved_cam2moved_cam_extended")
    parser.add_argument("--gt_root",    default="CamxTime_eval/eval_gt_wan2.1_format")
    parser.add_argument("--metadata",   default="metadata.csv")
    parser.add_argument("--output_dir", default="results/camxtime_metrics")
    parser.add_argument("--modes",      nargs="+", default=MODES)
    parser.add_argument("--device",     default="cuda" if torch.cuda.is_available() else "cpu")
    args = parser.parse_args()

    pred_root  = Path(args.pred_root)
    gt_root    = Path(args.gt_root)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    modes = args.modes

    all_scenes = [Path(fn).stem for fn in pd.read_csv(args.metadata)["file_name"]]
    t_global_start = time.time()

    _banner("SpaceTimePilot  ·  CamxTime Evaluation")
    print(f"  device    : {args.device}")
    print(f"  pred_root : {pred_root}")
    print(f"  gt_root   : {gt_root}")
    print(f"  metadata  : {args.metadata}  ({len(all_scenes)} scenes)")
    print(f"  modes     : {modes}")
    print(f"  output_dir: {output_dir}")
    _sep()

    # Check which modes / scenes actually have files up front
    print("\nFile availability check:")
    widths = [14, 10, 10, 38]
    _table_header(["Mode", "Pred found", "GT found", "Missing pred (first 3)"], widths)
    modes_available = []
    for mode in modes:
        if mode not in MODE_META:
            continue
        display, gt_pattern = MODE_META[mode]
        pred_dir = pred_root / mode
        pred_found = sum(1 for s in all_scenes if (pred_dir / f"{s}_cam_00.mp4").exists())
        gt_found   = sum(1 for s in all_scenes if (gt_root / s / f"moving_{gt_pattern}.mp4").exists())
        missing    = [s for s in all_scenes if not (pred_dir / f"{s}_cam_00.mp4").exists()][:3]
        miss_str   = ", ".join(missing) if missing else "—"
        _table_row([display, f"{pred_found}/{len(all_scenes)}",
                    f"{gt_found}/{len(all_scenes)}", miss_str], widths)
        if pred_found > 0 and gt_found > 0:
            modes_available.append(mode)
    _sep()

    print(f"\n  Will evaluate: {modes_available}\n")

    print("Loading LPIPS (AlexNet) …", end=" ", flush=True)
    lpips_net = lpips.LPIPS(net="alex").to(args.device)
    lpips_net.eval()
    print("done\n")

    all_video_rows, all_json, summary = [], {}, {}
    wb = Workbook(); wb.remove(wb.active)

    # Per-mode running table column widths
    col_w = [12, 7, 7, 7, 8, 6]  # scene, psnr, ssim, lpips, sec, frames

    for mode_i, mode in enumerate(modes_available):
        display, gt_pattern = MODE_META[mode]
        pred_dir = pred_root / mode

        _sep("═")
        print(f"  [{mode_i+1}/{len(modes_available)}]  {display}  "
              f"(pred: {mode}/  gt: moving_{gt_pattern}.mp4)")
        _sep("─")
        _table_header(["Scene", "PSNR", "SSIM", "LPIPS", "sec", "frames"], col_w)

        mode_rows, mode_json = [], {}
        t_mode_start = time.time()

        pbar = tqdm(all_scenes, desc=f"  {display}", unit="scene",
                    ncols=_W, leave=True)
        for scene in pbar:
            pred_path = pred_dir / f"{scene}_cam_00.mp4"
            gt_path   = gt_root  / scene / f"moving_{gt_pattern}.mp4"

            if not pred_path.exists():
                tqdm.write(f"  SKIP {scene}: pred not found")
                continue
            if not gt_path.exists():
                tqdm.write(f"  SKIP {scene}: GT not found")
                continue

            t0 = time.time()
            m = compute_video_metrics(load_frames(gt_path), load_frames(pred_path),
                                      lpips_net, args.device)
            elapsed = time.time() - t0

            row = {"mode": mode, "scene": scene, **m}
            mode_rows.append(row)
            all_video_rows.append(row)
            mode_json[scene] = m

            # Print table row (outside tqdm bar)
            tqdm.write(
                f"  │ {scene:<10s}  │ {m['psnr']:>6.3f}  │ {m['ssim']:>6.4f}  "
                f"│ {m['lpips']:>6.4f}  │ {elapsed:>5.1f}s │ {m['num_frames']:>4d}  │"
            )

            # Update tqdm postfix with running averages
            pbar.set_postfix(
                PSNR=f"{np.mean([r['psnr']  for r in mode_rows]):.3f}",
                SSIM=f"{np.mean([r['ssim']  for r in mode_rows]):.4f}",
                LPIPS=f"{np.mean([r['lpips'] for r in mode_rows]):.4f}",
            )

        pbar.close()

        if not mode_rows:
            print("  (no results for this mode)\n")
            continue

        t_mode = time.time() - t_mode_start
        s = {
            "num_videos": len(mode_rows),
            "psnr_mean":  float(np.mean([r["psnr"]  for r in mode_rows])),
            "psnr_std":   float(np.std( [r["psnr"]  for r in mode_rows])),
            "ssim_mean":  float(np.mean([r["ssim"]  for r in mode_rows])),
            "ssim_std":   float(np.std( [r["ssim"]  for r in mode_rows])),
            "lpips_mean": float(np.mean([r["lpips"] for r in mode_rows])),
            "lpips_std":  float(np.std( [r["lpips"] for r in mode_rows])),
        }
        summary[mode] = s
        mode_json["__summary__"] = s
        all_json[mode] = mode_json
        add_mode_sheet(wb, mode, mode_rows)

        _sep("─")
        print(f"  {display} summary  ({len(mode_rows)} videos, {t_mode:.0f}s)")
        print(f"    PSNR  {s['psnr_mean']:>7.3f} ± {s['psnr_std']:.3f}")
        print(f"    SSIM  {s['ssim_mean']:>7.4f} ± {s['ssim_std']:.4f}")
        print(f"    LPIPS {s['lpips_mean']:>7.4f} ± {s['lpips_std']:.4f}")

        # Highlight best / worst scene by PSNR
        best  = max(mode_rows, key=lambda r: r["psnr"])
        worst = min(mode_rows, key=lambda r: r["psnr"])
        print(f"    best  {best['scene']:<12s}  PSNR {best['psnr']:.3f}")
        print(f"    worst {worst['scene']:<12s}  PSNR {worst['psnr']:.3f}")
        print()

    # ── Write outputs ──────────────────────────────────────────────────────────
    _sep("═")
    print("Writing outputs …")
    with open(output_dir / "metrics.json", "w") as f:
        json.dump(all_json, f, indent=2)
    print(f"  metrics.json       → {output_dir/'metrics.json'}")

    with open(output_dir / "per_video.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["mode","scene","num_frames","psnr","ssim","lpips"])
        w.writeheader()
        for r in all_video_rows:
            w.writerow({k: r[k] for k in w.fieldnames})
    print(f"  per_video.csv      → {output_dir/'per_video.csv'}")

    with open(output_dir / "summary.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["mode","label","num_videos",
                    "psnr_mean","psnr_std","ssim_mean","ssim_std","lpips_mean","lpips_std"])
        for mode in modes:
            if mode not in summary:
                continue
            s = summary[mode]
            w.writerow([mode, MODE_META[mode][0], s["num_videos"],
                        round(s["psnr_mean"],3), round(s["psnr_std"],3),
                        round(s["ssim_mean"],4), round(s["ssim_std"],4),
                        round(s["lpips_mean"],4), round(s["lpips_std"],4)])
    print(f"  summary.csv        → {output_dir/'summary.csv'}")

    add_summary_sheet(wb, summary, modes)
    wb.save(output_dir / "results.xlsx")
    print(f"  results.xlsx       → {output_dir/'results.xlsx'}")

    scenes_with_data = [s for s in all_scenes
                        if any(r["scene"] == s for r in all_video_rows)]
    active_modes = [m for m in modes if m in summary]

    # ── Final summary table ────────────────────────────────────────────────────
    t_total = time.time() - t_global_start
    _sep("═")
    print("FINAL SUMMARY")
    _sep("─")
    fw = [14, 9, 10, 9, 10, 9, 10]
    _table_header(["Mode", "PSNR", "±", "SSIM", "±", "LPIPS", "±"], fw)
    for mode in modes:
        if mode not in summary:
            continue
        s = summary[mode]
        _table_row([MODE_META[mode][0],
                    f"{s['psnr_mean']:.3f}",  f"{s['psnr_std']:.3f}",
                    f"{s['ssim_mean']:.4f}",  f"{s['ssim_std']:.4f}",
                    f"{s['lpips_mean']:.4f}", f"{s['lpips_std']:.4f}"], fw)
    _sep("─")
    print(f"\n  Total time : {t_total/60:.1f} min  ({t_total:.0f}s)")
    print(f"  Output dir : {output_dir}")
    _sep("═")


if __name__ == "__main__":
    main()
